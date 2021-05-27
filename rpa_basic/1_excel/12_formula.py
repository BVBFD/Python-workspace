import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1, 2, 3)"
ws["A3"] = "=AVERAGE(1, 2, 3)"

ws["A4"] = 10 
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save("sample_formula.xlsx") 