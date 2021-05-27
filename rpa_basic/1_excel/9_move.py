from openpyxl import load_workbook
from random import *

wb = load_workbook("sample.xlsx")
ws = wb.active
# 번호 영어 수학
# 번호 (국어) 영어 수학

ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어" # B1 셀에 '국어' 입력

ws.delete_cols(2)

# for i in ws.iter_cols(min_col=2, max_col=2, min_row=2):
#     for cell in i:
#         cell.value = str(randint(1,100))

# de_ko = ws.iter_cols(min_row=2, max_col=2, min_col=2)
# for i in de_ko:
#     for y in i:
#         y.value = ""

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols = -1)

wb.save("sample_korean.xlsx")
wb.close()